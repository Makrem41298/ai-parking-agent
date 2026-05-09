import os
import re
import csv
import json
import hashlib
from typing import List, Dict, Any, Optional

import fitz
import pdfplumber
import openpyxl

from langchain_core.documents import Document
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter


class DocumentProcessor:
    def __init__(self, chunk_size: int = 700, chunk_overlap: int = 100):
        self.text_splitter = RecursiveCharacterTextSplitter(
            separators=["\n\n", ".\n", "\n", ". ", " "],
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    # =====================================================
    # METADATA
    # =====================================================

    def _safe_metadata(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        safe = {}

        for key, value in metadata.items():
            if value is None:
                continue

            if isinstance(value, (str, int, float, bool)):
                safe[key] = value
            elif isinstance(value, list):
                safe[key] = " > ".join(map(str, value))
            else:
                safe[key] = str(value)

        return safe

    def _text_metadata(
        self,
        filename: str,
        source: str,
        chunk_index: int,
        page: Optional[int] = None,
        section: Optional[str] = None,
        title: Optional[str] = None,
    ) -> Dict[str, Any]:
        metadata = {
            "file": filename,
            "source": source,
            "type": "text",
            "chunk_index": chunk_index,
        }

        if page is not None:
            metadata["page"] = page

        if section and section != "unknown":
            metadata["section"] = section

        if title and title != "UNKNOWN":
            metadata["title"] = title

        return self._safe_metadata(metadata)

    def _table_metadata(
        self,
        filename: str,
        source: str,
        chunk_index: int,
        page: Optional[int] = None,
        sheet: Optional[str] = None,
        method: Optional[str] = None,
    ) -> Dict[str, Any]:
        metadata = {
            "file": filename,
            "source": source,
            "type": "table",
            "chunk_index": chunk_index,
        }

        if page is not None:
            metadata["page"] = page

        if sheet:
            metadata["sheet"] = sheet

        if method:
            metadata["method"] = method

        return self._safe_metadata(metadata)

    # =====================================================
    # PDF
    # =====================================================

    def process_pdf_folder(self, folder_path: str) -> List[Document]:
        all_docs = []

        pdf_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(".pdf")
        ]

        print(f"\n📂 Found {len(pdf_files)} PDFs")

        for pdf_file in pdf_files:
            path = os.path.join(folder_path, pdf_file)

            print(f"\n🚀 Processing: {pdf_file}")

            docs = self.process_single_pdf(path)

            text_count = len([doc for doc in docs if doc.metadata.get("type") == "text"])
            table_count = len([doc for doc in docs if doc.metadata.get("type") == "table"])

            print(f"   ✅ Text chunks: {text_count}")
            print(f"   📊 Table chunks: {table_count}")

            all_docs.extend(docs)

        print(f"\n🔥 TOTAL DOCUMENTS: {len(all_docs)}")

        return all_docs

    def process_single_pdf(self, pdf_path: str) -> List[Document]:
        text_docs = self._process_pdf_text(pdf_path)
        table_docs = self._process_pdf_tables(pdf_path)

        table_docs = self._deduplicate(text_docs, table_docs)

        return text_docs + table_docs

    def _process_pdf_text(self, pdf_path: str) -> List[Document]:
        loader = PyMuPDFLoader(pdf_path)
        pages = loader.load()

        docs = []
        filename = os.path.basename(pdf_path)
        chunk_index = 0

        for page_num, page in enumerate(pages, start=1):
            text = self._clean_text(page.page_content)

            if len(text) < 50:
                continue

            sections = self._split_sections(text)

            for section in sections:
                if self._is_noise(section):
                    continue

                section_id = self._get_section_id(section)

                if section_id != "unknown":
                    title = self._get_title(section)
                else:
                    title = "UNKNOWN"

                chunks = self.text_splitter.create_documents([section])

                for chunk in chunks:
                    if self._is_low_quality_chunk(chunk.page_content):
                        continue

                    chunk.metadata = self._text_metadata(
                        filename=filename,
                        source="pdf",
                        page=page_num,
                        section=section_id,
                        title=title,
                        chunk_index=chunk_index,
                    )

                    docs.append(chunk)
                    chunk_index += 1

        return docs

    def _process_pdf_tables(self, pdf_path: str) -> List[Document]:
        docs = []
        filename = os.path.basename(pdf_path)
        chunk_index = 0

        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    tables = page.extract_tables()

                    for table in tables:
                        if not table or len(table) < 2:
                            continue

                        table_text = "\n".join(
                            [
                                " | ".join(
                                    str(cell).strip() if cell else ""
                                    for cell in row
                                )
                                for row in table
                            ]
                        )

                        if len(table_text.strip()) < 30:
                            continue

                        docs.append(
                            Document(
                                page_content=table_text,
                                metadata=self._table_metadata(
                                    filename=filename,
                                    source="pdf",
                                    page=page_num,
                                    method="pdfplumber",
                                    chunk_index=chunk_index,
                                ),
                            )
                        )

                        chunk_index += 1

        except Exception as error:
            print(f"⚠️ pdfplumber failed: {error}")

        try:
            pdf_doc = fitz.open(pdf_path)

            for page_num, page in enumerate(pdf_doc, start=1):
                blocks = page.get_text("blocks")

                for block in blocks:
                    text = block[4]

                    if not self._looks_like_table(text):
                        continue

                    docs.append(
                        Document(
                            page_content=text,
                            metadata=self._table_metadata(
                                filename=filename,
                                source="pdf",
                                page=page_num,
                                method="fallback",
                                chunk_index=chunk_index,
                            ),
                        )
                    )

                    chunk_index += 1

            pdf_doc.close()

        except Exception as error:
            print(f"⚠️ fallback failed: {error}")

        return docs

    # =====================================================
    # TEXT FILES
    # =====================================================

    def process_text_folder(self, folder_path: str, extension: str = ".txt") -> List[Document]:
        all_docs = []

        text_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(extension.lower())
        ]

        print(f"\n📂 Found {len(text_files)} text files")

        for filename in text_files:
            filepath = os.path.join(folder_path, filename)

            with open(filepath, "r", encoding="utf-8") as file:
                raw_text = file.read()

            cleaned_text = self._clean_text(raw_text)

            if len(cleaned_text) < 50:
                print(f"   ⚠️ Skipping {filename} - too short")
                continue

            docs = self._process_plain_text(
                text=cleaned_text,
                filename=filename,
                source="text_file",
            )

            print(f"   ✅ {filename}: {len(docs)} chunks")

            all_docs.extend(docs)

        print(f"\n🔥 TOTAL TEXT CHUNKS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # WORD
    # =====================================================

    def process_word_folder(self, folder_path: str) -> List[Document]:
        from langchain_community.document_loaders import Docx2txtLoader

        all_docs = []

        word_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(".docx")
        ]

        print(f"\n📂 Found {len(word_files)} Word files")

        for filename in word_files:
            filepath = os.path.join(folder_path, filename)

            loader = Docx2txtLoader(filepath)
            raw_docs = loader.load()

            if not raw_docs:
                continue

            full_text = raw_docs[0].page_content
            cleaned_text = self._clean_text(full_text)

            if len(cleaned_text) < 50:
                continue

            docs = self._process_plain_text(
                text=cleaned_text,
                filename=filename,
                source="word_document",
            )

            print(f"   ✅ {filename}: {len(docs)} chunks")

            all_docs.extend(docs)

        print(f"\n🔥 TOTAL WORD CHUNKS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # POWERPOINT
    # =====================================================

    def process_powerpoint_folder(self, folder_path: str) -> List[Document]:
        from langchain_community.document_loaders import UnstructuredPowerPointLoader

        all_docs = []

        ppt_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(".pptx")
        ]

        print(f"\n📂 Found {len(ppt_files)} PowerPoint files")

        for filename in ppt_files:
            filepath = os.path.join(folder_path, filename)

            loader = UnstructuredPowerPointLoader(filepath)
            data = loader.load()

            if not data:
                continue

            full_text = data[0].page_content
            cleaned_text = self._clean_text(full_text)

            if len(cleaned_text) < 20:
                continue

            docs = self._process_plain_text(
                text=cleaned_text,
                filename=filename,
                source="powerpoint",
            )

            print(f"   ✅ {filename}: {len(docs)} chunks")

            all_docs.extend(docs)

        print(f"\n🔥 TOTAL POWERPOINT CHUNKS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # EXCEL
    # =====================================================

    def process_excel_folder(
        self,
        folder_path: str,
        chunk_size: int = 0,
        chunk_overlap: int = 0,
    ) -> List[Document]:
        all_docs = []

        try:
            import xlrd
        except ImportError:
            xlrd = None

        excel_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith((".xlsx", ".xls"))
        ]

        print(f"\n📂 Found {len(excel_files)} Excel files")

        for filename in excel_files:
            filepath = os.path.join(folder_path, filename)
            file_docs = []

            if filename.lower().endswith(".xlsx"):
                workbook = openpyxl.load_workbook(filepath, data_only=True)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = [
                            str(cell).strip() if cell is not None else ""
                            for cell in row
                        ]
                        rows.append(cells)

                    file_docs.extend(
                        self._rows_to_table_documents(
                            rows=rows,
                            filename=filename,
                            source="excel",
                            sheet_name=sheet_name,
                            chunk_size=chunk_size,
                            chunk_overlap=chunk_overlap,
                        )
                    )

            elif filename.lower().endswith(".xls"):
                if xlrd is None:
                    print(f"   ⚠️ Skipping {filename} - xlrd not installed")
                    continue

                workbook = xlrd.open_workbook(filepath)

                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)

                    rows = []
                    for row_idx in range(sheet.nrows):
                        cells = [
                            str(sheet.cell_value(row_idx, col_idx)).strip()
                            for col_idx in range(sheet.ncols)
                        ]
                        rows.append(cells)

                    file_docs.extend(
                        self._rows_to_table_documents(
                            rows=rows,
                            filename=filename,
                            source="excel",
                            sheet_name=sheet_name,
                            chunk_size=chunk_size,
                            chunk_overlap=chunk_overlap,
                        )
                    )

            print(f"   ✅ {filename}: {len(file_docs)} chunks")

            all_docs.extend(file_docs)

        print(f"\n🔥 TOTAL EXCEL DOCUMENTS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # CSV
    # =====================================================

    def process_csv_folder(
        self,
        folder_path: str,
        delimiter: str = ",",
        chunk_size: int = 0,
        chunk_overlap: int = 0,
    ) -> List[Document]:
        all_docs = []

        csv_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(".csv")
        ]

        print(f"\n📂 Found {len(csv_files)} CSV files")

        for filename in csv_files:
            filepath = os.path.join(folder_path, filename)

            rows = []

            with open(filepath, "r", encoding="utf-8-sig") as file:
                reader = csv.reader(file, delimiter=delimiter)

                for row in reader:
                    rows.append([str(cell).strip() for cell in row])

            file_docs = self._rows_to_table_documents(
                rows=rows,
                filename=filename,
                source="csv",
                chunk_size=chunk_size,
                chunk_overlap=chunk_overlap,
            )

            print(f"   ✅ {filename}: {len(file_docs)} chunks")

            all_docs.extend(file_docs)

        print(f"\n🔥 TOTAL CSV DOCUMENTS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # JSON
    # =====================================================

    def process_json_folder(
        self,
        folder_path: str,
        mode: str = "auto",
        chunk_size: int = 0,
        chunk_overlap: int = 0,
        text_chunk_size: int = 0,
        text_chunk_overlap: int = 100,
    ) -> List[Document]:
        all_docs = []

        json_files = [
            file for file in os.listdir(folder_path)
            if file.lower().endswith(".json")
        ]

        print(f"\n📂 Found {len(json_files)} JSON files")

        for filename in json_files:
            filepath = os.path.join(folder_path, filename)

            with open(filepath, "r", encoding="utf-8") as file:
                data = json.load(file)

            if mode == "auto":
                file_mode = (
                    "table"
                    if isinstance(data, list) and all(isinstance(item, dict) for item in data)
                    else "text"
                )
            else:
                file_mode = mode

            file_docs = []

            if file_mode == "table":
                file_docs = self._json_table_to_documents(
                    data=data,
                    filename=filename,
                    chunk_size=chunk_size,
                    chunk_overlap=chunk_overlap,
                )

            else:
                try:
                    text_content = json.dumps(data, indent=2, ensure_ascii=False)
                except Exception:
                    text_content = str(data)

                if text_chunk_size > 0:
                    splitter = RecursiveCharacterTextSplitter(
                        separators=["\n\n", ".\n", "\n", ". ", " "],
                        chunk_size=text_chunk_size,
                        chunk_overlap=text_chunk_overlap,
                    )

                    chunks = splitter.create_documents([text_content])

                    for index, chunk in enumerate(chunks):
                        chunk.metadata = self._text_metadata(
                            filename=filename,
                            source="json",
                            chunk_index=index,
                        )
                        file_docs.append(chunk)

                else:
                    file_docs.append(
                        Document(
                            page_content=text_content,
                            metadata=self._text_metadata(
                                filename=filename,
                                source="json",
                                chunk_index=0,
                            ),
                        )
                    )

            print(f"   ✅ {filename}: {len(file_docs)} chunks mode={file_mode}")

            all_docs.extend(file_docs)

        print(f"\n🔥 TOTAL JSON CHUNKS: {len(all_docs)}")

        return all_docs

    # =====================================================
    # COMMON TEXT PROCESSOR
    # =====================================================

    def _process_plain_text(
        self,
        text: str,
        filename: str,
        source: str,
    ) -> List[Document]:
        docs = []

        sections = self._split_sections(text)
        chunk_index = 0

        for section in sections:
            if self._is_noise(section):
                continue

            section_id = self._get_section_id(section)

            if section_id != "unknown":
                title = self._get_title(section)
            else:
                title = "UNKNOWN"

            chunks = self.text_splitter.create_documents([section])

            for chunk in chunks:
                if self._is_low_quality_chunk(chunk.page_content):
                    continue

                chunk.metadata = self._text_metadata(
                    filename=filename,
                    source=source,
                    section=section_id,
                    title=title,
                    chunk_index=chunk_index,
                )

                docs.append(chunk)
                chunk_index += 1

        return docs

    # =====================================================
    # TABLE HELPERS
    # =====================================================

    def _rows_to_table_documents(
        self,
        rows: List[List[str]],
        filename: str,
        source: str,
        sheet_name: Optional[str] = None,
        chunk_size: int = 0,
        chunk_overlap: int = 0,
    ) -> List[Document]:
        docs = []

        if len(rows) < 2:
            return docs

        header = rows[0]
        data_rows = rows[1:]

        if chunk_size <= 0 or len(data_rows) <= chunk_size:
            table_text = "\n".join([" | ".join(row) for row in rows])

            if table_text.strip():
                docs.append(
                    Document(
                        page_content=table_text,
                        metadata=self._table_metadata(
                            filename=filename,
                            source=source,
                            sheet=sheet_name,
                            chunk_index=0,
                        ),
                    )
                )

            return docs

        step = max(chunk_size - chunk_overlap, 1)

        for start in range(0, len(data_rows), step):
            chunk_rows = data_rows[start:start + chunk_size]
            table_rows = [header] + chunk_rows

            table_text = "\n".join([" | ".join(row) for row in table_rows])

            docs.append(
                Document(
                    page_content=table_text,
                    metadata=self._table_metadata(
                        filename=filename,
                        source=source,
                        sheet=sheet_name,
                        chunk_index=len(docs),
                    ),
                )
            )

        return docs

    def _json_table_to_documents(
        self,
        data: Any,
        filename: str,
        chunk_size: int = 0,
        chunk_overlap: int = 0,
    ) -> List[Document]:
        if not isinstance(data, list) or not data:
            return []

        if not all(isinstance(item, dict) for item in data):
            return []

        keys = list(data[0].keys())

        rows = [[str(key) for key in keys]]

        for item in data:
            rows.append(
                [
                    str(item.get(key, "")).strip()
                    for key in keys
                ]
            )

        return self._rows_to_table_documents(
            rows=rows,
            filename=filename,
            source="json",
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    # =====================================================
    # CLEANING
    # =====================================================

    def _clean_text(self, text: str) -> str:
        if not text:
            return ""

        # Fix broken words: "trans-\nport" => "transport"
        text = re.sub(r"-\n", "", text)

        # Normalize bullets
        text = text.replace("•", "\n- ")
        text = text.replace("·", "\n- ")
        text = text.replace("+", "\n- ")

        # Remove image placeholder text
        text = re.sub(
            r"(?i)(image placeholder|file:\s*\S+\.(png|jpg|jpeg|gif|webp))",
            "",
            text,
        )

        # Remove repeated spaces
        text = re.sub(r"[ \t]+", " ", text)

        # Remove standalone page numbers
        text = re.sub(r"^\s*\d{1,3}\s*$", "", text, flags=re.MULTILINE)

        # Add newline before headings like "2.3 Solution envisagée"
        text = re.sub(
            r"(?<!\n)(\b\d+(\.\d+)+\.?\s+[A-ZÀ-Ý])",
            r"\n\1",
            text,
        )

        # Add newline before headings like "3 Spécifications des besoins"
        text = re.sub(
            r"(?<!\n)(\b\d+\.?\s+[A-ZÀ-Ý][A-Za-zÀ-ÿ])",
            r"\n\1",
            text,
        )

        # Add newline before "Chapter 1" / "Chapitre 1"
        text = re.sub(
            r"(?<!\n)((Chapter|Chapitre|CHAPITRE)\s+\d+)",
            r"\n\1",
            text,
            flags=re.IGNORECASE,
        )

        # Clean many newlines
        text = re.sub(r"\n{3,}", "\n\n", text)

        return text.strip()

    # =====================================================
    # SECTION DETECTION
    # =====================================================

    def _split_sections(self, text: str) -> List[str]:
        pattern = r"""
            (?=
                \n
                (?:
                    Chapter\s+\d+
                    |
                    CHAPITRE\s+\d+
                    |
                    Chapitre\s+\d+
                    |
                    \d+\.\d+\.\d+\.?\s+[A-ZÀ-Ý]
                    |
                    \d+\.\d+\.?\s+[A-ZÀ-Ý]
                    |
                    \d+\.?\s+[A-ZÀ-Ý]
                )
            )
        """

        sections = re.split(pattern, text, flags=re.VERBOSE | re.IGNORECASE)

        return [section.strip() for section in sections if section.strip()]

    def _get_section_id(self, text: str) -> str:
        lines = [line.strip() for line in text.split("\n") if line.strip()]

        if not lines:
            return "unknown"

        first_line = lines[0]
        first_lower = first_line.lower()

        ignored_titles = [
            "résumé",
            "resume",
            "abstract",
            "bibliographie",
            "references",
            "table des matières",
            "table des figures",
            "liste des tableaux",
            "conclusion générale",
            "introduction générale",
        ]

        if first_lower in ignored_titles:
            return "unknown"

        chapter_match = re.match(
            r"^(Chapter|CHAPITRE|Chapitre)\s+(\d+)",
            first_line,
            re.IGNORECASE,
        )

        if chapter_match:
            return f"{chapter_match.group(1)} {chapter_match.group(2)}"

        # Example: 2.3 Solution envisagée
        section_match = re.match(
            r"^(\d+(\.\d+)+)\.?\s+[A-ZÀ-Ý]",
            first_line,
        )

        if section_match:
            return section_match.group(1)

        # Example: 3 Spécifications des besoins
        main_match = re.match(
            r"^(\d+)\.?\s+[A-ZÀ-Ý]",
            first_line,
        )

        if main_match:
            number = int(main_match.group(1))

            # Avoid false sections like 76, 79, 132
            if 1 <= number <= 20:
                return str(number)

        # Example:
        # 2.3
        # Solution envisagée
        if re.match(r"^\d+(\.\d+)+$", first_line):
            return first_line

        # Example:
        # 3
        # Spécifications des besoins
        if re.match(r"^\d+$", first_line):
            number = int(first_line)

            if 1 <= number <= 20:
                return first_line

        return "unknown"

    def _get_title(self, text: str) -> str:
        lines = [line.strip() for line in text.split("\n") if line.strip()]

        if not lines:
            return "UNKNOWN"

        first_line = lines[0]

        ignored_titles = [
            "résumé",
            "resume",
            "abstract",
            "bibliographie",
            "references",
            "table des matières",
            "table des figures",
            "liste des tableaux",
        ]

        if first_line.lower() in ignored_titles:
            return "UNKNOWN"

        # Example: 2.3 Solution envisagée
        match = re.match(r"^\d+(\.\d+)*\.?\s+(.+)", first_line)

        if match:
            title = match.group(2).strip()
            return self._short_title(title)

        # Example:
        # 2.3
        # Solution envisagée
        if re.match(r"^\d+(\.\d+)*$", first_line) and len(lines) > 1:
            return self._short_title(lines[1])

        # Example:
        # Chapter 1
        # Implementation and Realization
        if re.match(r"^(Chapter|CHAPITRE|Chapitre)\s+\d+", first_line, re.IGNORECASE):
            if len(lines) > 1:
                return self._short_title(lines[1])

            return self._short_title(first_line)

        return self._short_title(first_line)

    def _short_title(self, title: str, max_words: int = 8) -> str:
        title = title.strip()
        title = re.sub(r"\s+", " ", title)

        stop_patterns = [
            r"\s+This\s+",
            r"\s+The\s+",
            r"\s+To\s+",
            r"\s+Users\s+",
            r"\s+If\s+",
            r"\s+For\s+",
            r"\s+Dans\s+",
            r"\s+Notre\s+",
            r"\s+Le\s+",
            r"\s+La\s+",
            r"\s+Les\s+",
            r"\s+Ce\s+",
            r"\s+Cette\s+",
        ]

        for pattern in stop_patterns:
            match = re.search(pattern, title)
            if match and match.start() > 3:
                title = title[:match.start()].strip()
                break

        words = title.split()

        if len(words) > max_words:
            title = " ".join(words[:max_words])

        return title or "UNKNOWN"

    # =====================================================
    # FILTERS
    # =====================================================

    def _is_noise(self, text: str) -> bool:
        content = text.strip()
        lower_content = content.lower()

        if len(content) < 50:
            return True

        noise_titles = [
            "table des matières",
            "table des figures",
            "liste des tableaux",
        ]

        if any(title in lower_content for title in noise_titles):
            return True

        if "contents" in lower_content and re.search(r"\.{3,}", content):
            return True

        if re.search(r"(\.\s*){8,}", content):
            return True

        alpha_chars = sum(1 for char in content if char.isalpha())
        total_chars = max(len(content), 1)

        if alpha_chars / total_chars < 0.25:
            return True

        return False

    def _is_low_quality_chunk(self, text: str) -> bool:
        content = text.strip()

        if len(content) < 40:
            return True

        alpha_chars = sum(1 for char in content if char.isalpha())
        total_chars = max(len(content), 1)

        if alpha_chars / total_chars < 0.25:
            return True

        if re.search(r"(\.\s*){8,}", content):
            return True

        return False

    def _looks_like_table(self, text: str) -> bool:
        if not text or len(text.strip()) < 30:
            return False

        lines = [line for line in text.split("\n") if line.strip()]

        if len(lines) < 3:
            return False

        pipe_lines = [line for line in lines if "|" in line]

        if len(pipe_lines) >= 3:
            column_counts = [line.count("|") for line in pipe_lines]

            if len(set(column_counts)) <= 2:
                return True

        return False

    # =====================================================
    # DEDUPLICATION
    # =====================================================

    def _deduplicate(
        self,
        text_docs: List[Document],
        table_docs: List[Document],
    ) -> List[Document]:
        if not text_docs or not table_docs:
            return table_docs

        text_fingerprints = {
            self._fingerprint(doc.page_content)
            for doc in text_docs
        }

        filtered_docs = []

        for doc in table_docs:
            fingerprint = self._fingerprint(doc.page_content)

            if fingerprint not in text_fingerprints:
                filtered_docs.append(doc)

        return filtered_docs

    def _fingerprint(self, text: str) -> str:
        cleaned = re.sub(r"\s+", "", text.lower())[:250]

        return hashlib.md5(cleaned.encode()).hexdigest()

    # =====================================================
    # STABLE IDS FOR CHROMA
    # =====================================================

    def make_chroma_ids(self, docs: List[Document]) -> List[str]:
        ids = []

        for doc in docs:
            file_name = doc.metadata.get("file", "unknown")
            page = doc.metadata.get("page", 0)
            chunk_index = doc.metadata.get("chunk_index", 0)

            raw_id = f"{file_name}-{page}-{chunk_index}-{doc.page_content[:80]}"
            stable_id = hashlib.md5(raw_id.encode()).hexdigest()

            ids.append(stable_id)

        return ids