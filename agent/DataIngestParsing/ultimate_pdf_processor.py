import os, re, uuid, hashlib, fitz, pdfplumber
from typing import List, Dict
from langchain_core.documents import Document
from langchain_community.document_loaders import PyMuPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter


class UltimatePDFProcessor:

    def __init__(self, chunk_size=700, chunk_overlap=100):
        self.text_splitter = RecursiveCharacterTextSplitter(
            separators=[".\n", "\n\n", "\n", ". ", " "],
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )

    # ── ENTRY: PDF FOLDER ──
    def process_folder(self, folder_path: str) -> List[Document]:
        all_docs = []
        pdf_files = [f for f in os.listdir(folder_path) if f.endswith(".pdf")]
        print(f"\n📂 Found {len(pdf_files)} PDFs")
        for pdf in pdf_files:
            path = os.path.join(folder_path, pdf)
            print(f"\n🚀 Processing: {pdf}")
            docs = self.process_single_pdf(path)
            txt = len([d for d in docs if d.metadata['type'] == 'text'])
            tbl = len([d for d in docs if d.metadata['type'] == 'table'])
            print(f"   ✅ Text chunks: {txt}")
            print(f"   📊 Table chunks: {tbl}")
            all_docs.extend(docs)
        print(f"\n🔥 TOTAL DOCUMENTS: {len(all_docs)}")
        return all_docs

    def process_single_pdf(self, pdf_path: str) -> List[Document]:
        text_docs = self._process_text(pdf_path)
        table_docs = self._process_tables(pdf_path)
        # Deduplicate: remove table docs that overlap with text docs
        table_docs = self._deduplicate(text_docs, table_docs)
        return text_docs + table_docs

    # ── TEXT PROCESSING ──
    def _process_text(self, pdf_path: str) -> List[Document]:
        loader = PyMuPDFLoader(pdf_path)
        pages = loader.load()
        docs = []
        document_id = str(uuid.uuid4())
        chunk_index = 0
        current_section = "unknown"
        current_title = "UNKNOWN"
        for page_num, page in enumerate(pages):
            text = self._clean_text(page.page_content)
            if len(text) < 50:
                continue
            sections = self._split_sections(text)
            for section in sections:
                if self._is_noise(section):
                    continue
                sid = self._get_section_id(section)
                title = self._get_title(section)
                if sid != "unknown":
                    current_section = sid
                    current_title = title
                hierarchy = self._get_hierarchy(current_section, current_title)
                chunks = self.text_splitter.create_documents(
                    texts=[section],
                    metadatas=[{
                        "file": os.path.basename(pdf_path),
                        "page": page_num + 1,
                        "type": "text",
                        "section": current_section,
                        "title": current_title,
                        "hierarchy": hierarchy,
                        "document_id": document_id
                    }]
                )
                for c in chunks:
                    # Post-validate chunk quality
                    if self._is_low_quality_chunk(c.page_content):
                        continue
                    c.metadata["chunk_id"] = str(uuid.uuid4())
                    c.metadata["chunk_index"] = chunk_index
                    chunk_index += 1
                    docs.append(c)
        return docs

    # ── TABLE PROCESSING ──
    def _process_tables(self, pdf_path: str) -> List[Document]:
        docs = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for i, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        text_table = "\n".join(
                            [" | ".join([str(cell) if cell else "" for cell in row]) for row in table]
                        )
                        if len(text_table.strip()) < 30:
                            continue
                        docs.append(Document(
                            page_content=text_table,
                            metadata={
                                "file": os.path.basename(pdf_path),
                                "page": i + 1,
                                "type": "table",
                                "method": "pdfplumber"
                            }
                        ))
        except Exception as e:
            print("⚠️ pdfplumber failed:", e)
        # Fallback with STRICT table detection
        try:
            doc = fitz.open(pdf_path)
            for page_num, page in enumerate(doc):
                blocks = page.get_text("blocks")
                for b in blocks:
                    text = b[4]
                    if self._looks_like_table(text):
                        docs.append(Document(
                            page_content=text,
                            metadata={
                                "file": os.path.basename(pdf_path),
                                "page": page_num + 1,
                                "type": "table",
                                "method": "fallback"
                            }
                        ))
        except Exception as e:
            print("⚠️ fallback failed:", e)
        return docs

    # ── IMPROVED: CLEAN TEXT ──
    def _clean_text(self, text: str) -> str:
        text = re.sub(r'-\n', '', text)
        text = re.sub(r'[ \t]+', ' ', text)
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Remove standalone page numbers at end of pages
        text = re.sub(r'\n\d{1,3}\s*$', '', text)
        # Remove standalone page numbers on their own line
        text = re.sub(r'^\d{1,3}\s*$', '', text, flags=re.MULTILINE)
        # Remove image placeholder noise
        text = re.sub(r'(?i)(?:image placeholder|file:\s*\S+\.(?:png|jpg|jpeg|gif|PNG))', '', text)
        return text.strip()

    # ── IMPROVED: SECTION SPLITTING ──
    def _split_sections(self, text: str) -> List[str]:
        pattern = r'''
            (?=
                \n
                (?:
                    Chapter\s+\d+\s*\n
                    |
                    \d+\.\d+\.\d+\s*\n
                    |
                    \d+\.\d+\s*\n
                    |
                    \d+\s*\n
                )
            )
        '''
        sections = re.split(pattern, text, flags=re.VERBOSE)
        return [s.strip() for s in sections if s.strip()]

    # ── IMPROVED: TITLE EXTRACTION ──
    def _get_title(self, text: str) -> str:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if not lines:
            return "UNKNOWN"
        if re.match(r'^\d+(\.\d+)*$', lines[0]) and len(lines) > 1:
            return lines[1]
        match = re.match(r'^\d+(\.\d+)*\s+(.*)', lines[0])
        if match:
            return match.group(2)
        if lines[0].startswith("Chapter") and len(lines) > 1:
            return lines[1]
        return lines[0]

    # ── IMPROVED: SECTION ID (strict) ──
    def _get_section_id(self, text: str) -> str:
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if not lines:
            return "unknown"
        first = lines[0]
        # Chapter heading
        if re.match(r'^Chapter\s+\d+$', first):
            return first
        # Numbered section like "1.2.3" → valid
        if re.match(r'^\d+(\.\d+)+$', first):
            return first
        # Single number → only valid if next line looks like a real title
        if re.match(r'^\d+$', first):
            if len(lines) > 1:
                next_line = lines[1]
                # Real title: mostly alpha, >3 chars, starts uppercase
                if (len(next_line) > 3 and
                    next_line[0].isupper() and
                    sum(c.isalpha() or c.isspace() for c in next_line) / max(len(next_line), 1) > 0.7):
                    return first
            return "unknown"
        return "unknown"

    def _get_hierarchy(self, section_id: str, title: str) -> Dict:
        if section_id == "unknown":
            return {"level": 0, "path": [title]}
        return {
            "level": len(section_id.split(".")),
            "path": [section_id, title]
        }

    # ── IMPROVED: NOISE FILTER ──
    def _is_noise(self, text: str) -> bool:
        t = text.strip()
        lower_t = t.lower()
        if len(t) < 50:
            return True
        if "contents" in lower_t and re.search(r'\.{3,}', lower_t):
            return True
        # Spaced dots → ToC lines
        if re.search(r'(\.\s){4,}', t):
            return True
        # Mostly dots
        dot_count = t.count('.') + t.count('. ')
        if dot_count > len(t) * 0.3:
            return True
        # Just a page number
        if re.match(r'^\d{1,3}$', t):
            return True
        return False

    # ── IMPROVED: TABLE DETECTOR (strict) ──
    def _looks_like_table(self, text: str) -> bool:
        if not text or len(text.strip()) < 30:
            return False
        lines = [l for l in text.split("\n") if l.strip()]
        if len(lines) < 3:
            return False
        pipe_lines = [l for l in lines if "|" in l]
        # Need at least 3 lines with pipes
        if len(pipe_lines) >= 3:
            col_counts = [l.count("|") for l in pipe_lines]
            if len(set(col_counts)) <= 2:
                return True
        return False

    # ── NEW: LOW QUALITY CHUNK FILTER ──
    def _is_low_quality_chunk(self, text: str) -> bool:
        t = text.strip()
        if len(t) < 40:
            return True
        # Mostly whitespace/dots
        real_chars = sum(1 for c in t if c.isalnum())
        if real_chars < len(t) * 0.3:
            return True
        # ToC-like content
        if re.search(r'(\.\s){4,}', t):
            return True
        return False

    # ── NEW: DEDUPLICATION ──
    def _deduplicate(self, text_docs: List[Document], table_docs: List[Document]) -> List[Document]:
        if not text_docs or not table_docs:
            return table_docs
        text_fingerprints = set()
        for doc in text_docs:
            fp = self._fingerprint(doc.page_content)
            text_fingerprints.add(fp)
        filtered = []
        for doc in table_docs:
            fp = self._fingerprint(doc.page_content)
            if fp not in text_fingerprints:
                filtered.append(doc)
        return filtered

    def _fingerprint(self, text: str) -> str:
        cleaned = re.sub(r'\s+', '', text.lower())[:200]
        return hashlib.md5(cleaned.encode()).hexdigest()

    # ── TEXT FILES ──
    def process_text_folder(self, folder_path: str, extension=".txt") -> List[Document]:
        all_docs = []
        text_files = [f for f in os.listdir(folder_path) if f.endswith(extension)]
        print(f"\n📂 Found {len(text_files)} text files in {folder_path}")
        for filename in text_files:
            filepath = os.path.join(folder_path, filename)
            with open(filepath, "r", encoding="utf-8") as f:
                raw_text = f.read()
            cleaned = self._clean_text(raw_text)
            if len(cleaned) < 50:
                print(f"   ⚠️ Skipping {filename} – too short")
                continue
            sections = self._split_sections(cleaned)
            document_id = str(uuid.uuid4())
            chunk_index = 0
            current_section = "unknown"
            current_title = "UNKNOWN"
            file_chunks_count = 0
            for section in sections:
                if self._is_noise(section):
                    continue
                sid = self._get_section_id(section)
                title = self._get_title(section)
                if sid != "unknown":
                    current_section = sid
                    current_title = title
                hierarchy = self._get_hierarchy(current_section, current_title)
                chunks = self.text_splitter.create_documents(
                    texts=[section],
                    metadatas=[{
                        "file": filename, "type": "text", "source": "text_file",
                        "section": current_section, "title": current_title,
                        "hierarchy": hierarchy, "document_id": document_id
                    }]
                )
                for c in chunks:
                    if self._is_low_quality_chunk(c.page_content):
                        continue
                    c.metadata["chunk_id"] = str(uuid.uuid4())
                    c.metadata["chunk_index"] = chunk_index
                    chunk_index += 1
                    file_chunks_count += 1
                    all_docs.append(c)
            print(f"   ✅ {filename}: {file_chunks_count} chunks total")
        print(f"\n🔥 TOTAL TEXT CHUNKS: {len(all_docs)}")
        return all_docs

    # ── WORD DOCUMENTS ──
    def process_word_folder(self, folder_path: str) -> List[Document]:
        from langchain_community.document_loaders import Docx2txtLoader
        all_docs = []
        word_files = [f for f in os.listdir(folder_path) if f.endswith(".docx")]
        print(f"\n📂 Found {len(word_files)} Word files in {folder_path}")
        for filename in word_files:
            filepath = os.path.join(folder_path, filename)
            loader = Docx2txtLoader(filepath)
            raw_docs = loader.load()
            if not raw_docs:
                continue
            full_text = raw_docs[0].page_content
            if not full_text or len(full_text.strip()) < 50:
                continue
            cleaned = self._clean_text(full_text)
            sections = self._split_sections(cleaned)
            document_id = str(uuid.uuid4())
            chunk_index = 0
            current_section = "unknown"
            current_title = "UNKNOWN"
            for section in sections:
                if self._is_noise(section):
                    continue
                sid = self._get_section_id(section)
                title = self._get_title(section)
                if sid != "unknown":
                    current_section = sid
                    current_title = title
                hierarchy = self._get_hierarchy(current_section, current_title)
                chunks = self.text_splitter.create_documents(
                    texts=[section],
                    metadatas=[{
                        "file": filename, "type": "text", "source": "word_document",
                        "section": current_section, "title": current_title,
                        "hierarchy": hierarchy, "document_id": document_id
                    }]
                )
                for c in chunks:
                    if self._is_low_quality_chunk(c.page_content):
                        continue
                    c.metadata["chunk_id"] = str(uuid.uuid4())
                    c.metadata["chunk_index"] = chunk_index
                    chunk_index += 1
                    all_docs.append(c)
            print(f"   ✅ {filename}: {chunk_index} chunks")
        print(f"\n🔥 TOTAL WORD CHUNKS: {len(all_docs)}")
        return all_docs

    # ── POWERPOINT ──
    def process_powerpoint_folder(self, folder_path: str) -> List[Document]:
        from langchain_community.document_loaders import UnstructuredPowerPointLoader
        all_docs = []
        ppt_files = [f for f in os.listdir(folder_path) if f.endswith(".pptx")]
        print(f"\n📂 Found {len(ppt_files)} PowerPoint files in {folder_path}")
        for filename in ppt_files:
            filepath = os.path.join(folder_path, filename)
            loader = UnstructuredPowerPointLoader(filepath)
            data = loader.load()
            if not data:
                continue
            full_text = data[0].page_content
            if len(full_text.strip()) < 20:
                continue
            cleaned = self._clean_text(full_text)
            sections = self._split_sections(cleaned)
            document_id = str(uuid.uuid4())
            chunk_index = 0
            current_section = "unknown"
            current_title = "UNKNOWN"
            for section in sections:
                if self._is_noise(section):
                    continue
                sid = self._get_section_id(section)
                title = self._get_title(section)
                if sid != "unknown":
                    current_section = sid
                    current_title = title
                hierarchy = self._get_hierarchy(current_section, current_title)
                chunks = self.text_splitter.create_documents(
                    texts=[section],
                    metadatas=[{
                        "file": filename, "type": "text", "source": "powerpoint",
                        "section": current_section, "title": current_title,
                        "hierarchy": hierarchy, "document_id": document_id
                    }]
                )
                for c in chunks:
                    if self._is_low_quality_chunk(c.page_content):
                        continue
                    c.metadata["chunk_id"] = str(uuid.uuid4())
                    c.metadata["chunk_index"] = chunk_index
                    chunk_index += 1
                    all_docs.append(c)
            print(f"   ✅ {filename}: {chunk_index} chunks")
        print(f"\n🔥 TOTAL POWERPOINT CHUNKS: {len(all_docs)}")
        return all_docs

    # ── EXCEL ──
    def process_excel_folder(self, folder_path: str, chunk_size=0, chunk_overlap=0) -> List[Document]:
        all_docs = []
        import openpyxl
        try:
            import xlrd
        except ImportError:
            xlrd = None
        excel_files = [f for f in os.listdir(folder_path) if f.endswith((".xlsx", ".xls"))]
        print(f"\n📂 Found {len(excel_files)} Excel files in {folder_path}")
        for filename in excel_files:
            filepath = os.path.join(folder_path, filename)
            if filename.endswith(".xlsx"):
                workbook = openpyxl.load_workbook(filepath, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(cell) if cell is not None else "" for cell in row]
                        rows.append(cells)
                    if not rows:
                        continue
                    header = rows[0]
                    data_rows = rows[1:]
                    if chunk_size <= 0 or len(data_rows) <= chunk_size:
                        table_text = "\n".join([" | ".join(r) for r in rows])
                        if table_text.strip():
                            all_docs.append(Document(page_content=table_text, metadata={
                                "file": filename, "type": "table", "source": "excel",
                                "sheet": sheet_name, "document_id": str(uuid.uuid4()),
                                "chunk_id": str(uuid.uuid4()), "chunk_index": len(all_docs)
                            }))
                    else:
                        step = chunk_size - chunk_overlap
                        for start in range(0, len(data_rows), step):
                            chunk_rows = data_rows[start:start + chunk_size]
                            table_text = "\n".join([" | ".join(r) for r in [header] + chunk_rows])
                            all_docs.append(Document(page_content=table_text, metadata={
                                "file": filename, "type": "table", "source": "excel",
                                "sheet": sheet_name, "document_id": str(uuid.uuid4()),
                                "chunk_id": str(uuid.uuid4()), "chunk_index": len(all_docs)
                            }))
            elif filename.endswith(".xls"):
                if xlrd is None:
                    print(f"   ⚠️ Skipping {filename} – xlrd not installed")
                    continue
                workbook = xlrd.open_workbook(filepath)
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    rows = []
                    for row_idx in range(sheet.nrows):
                        cells = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                        rows.append(cells)
                    if not rows:
                        continue
                    header = rows[0]
                    data_rows = rows[1:]
                    if chunk_size <= 0 or len(data_rows) <= chunk_size:
                        table_text = "\n".join([" | ".join(r) for r in rows])
                        if table_text.strip():
                            all_docs.append(Document(page_content=table_text, metadata={
                                "file": filename, "type": "table", "source": "excel",
                                "sheet": sheet_name, "document_id": str(uuid.uuid4()),
                                "chunk_id": str(uuid.uuid4()), "chunk_index": len(all_docs)
                            }))
                    else:
                        step = chunk_size - chunk_overlap
                        for start in range(0, len(data_rows), step):
                            chunk_rows = data_rows[start:start + chunk_size]
                            table_text = "\n".join([" | ".join(r) for r in [header] + chunk_rows])
                            all_docs.append(Document(page_content=table_text, metadata={
                                "file": filename, "type": "table", "source": "excel",
                                "sheet": sheet_name, "document_id": str(uuid.uuid4()),
                                "chunk_id": str(uuid.uuid4()), "chunk_index": len(all_docs)
                            }))
            print(f"   ✅ {filename}: {len(all_docs)} chunks total")
        print(f"\n🔥 TOTAL EXCEL DOCUMENTS: {len(all_docs)}")
        return all_docs

    # ── CSV ──
    def process_csv_folder(self, folder_path: str, delimiter=",", chunk_size=0, chunk_overlap=0) -> List[Document]:
        import csv
        all_docs = []
        csv_files = [f for f in os.listdir(folder_path) if f.endswith(".csv")]
        print(f"\n📂 Found {len(csv_files)} CSV files in {folder_path}")
        for filename in csv_files:
            filepath = os.path.join(folder_path, filename)
            rows = []
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    rows.append([str(cell).strip() for cell in row])
            if len(rows) < 2:
                continue
            header = rows[0]
            data_rows = rows[1:]
            if chunk_size <= 0 or len(data_rows) <= chunk_size:
                table_text = "\n".join([" | ".join(r) for r in rows])
                if table_text.strip():
                    all_docs.append(Document(page_content=table_text, metadata={
                        "file": filename, "type": "table", "source": "csv",
                        "document_id": str(uuid.uuid4()), "chunk_id": str(uuid.uuid4()),
                        "chunk_index": len(all_docs)
                    }))
            else:
                step = chunk_size - chunk_overlap
                for start in range(0, len(data_rows), step):
                    chunk_rows = data_rows[start:start + chunk_size]
                    table_text = "\n".join([" | ".join(r) for r in [header] + chunk_rows])
                    all_docs.append(Document(page_content=table_text, metadata={
                        "file": filename, "type": "table", "source": "csv",
                        "document_id": str(uuid.uuid4()), "chunk_id": str(uuid.uuid4()),
                        "chunk_index": len(all_docs)
                    }))
            print(f"   ✅ {filename}: {len(all_docs)} chunks total")
        print(f"\n🔥 TOTAL CSV DOCUMENTS: {len(all_docs)}")
        return all_docs

    # ── JSON ──
    def process_json_folder(self, folder_path: str, mode="auto", chunk_size=0,
                            chunk_overlap=0, text_chunk_size=0, text_chunk_overlap=100) -> List[Document]:
        import json
        all_docs = []
        json_files = [f for f in os.listdir(folder_path) if f.endswith(".json")]
        print(f"\n📂 Found {len(json_files)} JSON files in {folder_path}")
        for filename in json_files:
            filepath = os.path.join(folder_path, filename)
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if mode == "auto":
                mode_file = "table" if isinstance(data, list) and all(isinstance(i, dict) for i in data) else "text"
            else:
                mode_file = mode
            document_id = str(uuid.uuid4())
            if mode_file == "table":
                if not data:
                    continue
                keys = list(data[0].keys())
                header = [str(k) for k in keys]
                rows = [header] + [[str(obj.get(k, "")) for k in keys] for obj in data]
                if chunk_size <= 0 or len(rows) - 1 <= chunk_size:
                    table_text = "\n".join([" | ".join(r) for r in rows])
                    if table_text.strip():
                        all_docs.append(Document(page_content=table_text, metadata={
                            "file": filename, "type": "table", "source": "json",
                            "document_id": document_id, "chunk_id": str(uuid.uuid4()),
                            "chunk_index": len(all_docs)
                        }))
                else:
                    header_row = rows[0]
                    d_rows = rows[1:]
                    step = chunk_size - chunk_overlap
                    for start in range(0, len(d_rows), step):
                        cr = [header_row] + d_rows[start:start + chunk_size]
                        table_text = "\n".join([" | ".join(r) for r in cr])
                        all_docs.append(Document(page_content=table_text, metadata={
                            "file": filename, "type": "table", "source": "json",
                            "document_id": document_id, "chunk_id": str(uuid.uuid4()),
                            "chunk_index": len(all_docs)
                        }))
            else:
                try:
                    text_content = json.dumps(data, indent=2, ensure_ascii=False)
                except Exception:
                    text_content = str(data)
                if text_chunk_size > 0:
                    temp_splitter = RecursiveCharacterTextSplitter(
                        separators=[".\n", "\n\n", "\n", ". ", " "],
                        chunk_size=text_chunk_size, chunk_overlap=text_chunk_overlap
                    )
                    chunks = temp_splitter.create_documents(
                        texts=[text_content],
                        metadatas=[{"file": filename, "type": "text", "source": "json", "document_id": document_id}]
                    )
                    for idx, c in enumerate(chunks):
                        c.metadata["chunk_id"] = str(uuid.uuid4())
                        c.metadata["chunk_index"] = len(all_docs) + idx
                    all_docs.extend(chunks)
                else:
                    all_docs.append(Document(page_content=text_content, metadata={
                        "file": filename, "type": "text", "source": "json",
                        "document_id": document_id, "chunk_id": str(uuid.uuid4()),
                        "chunk_index": len(all_docs)
                    }))
            print(f"   ✅ {filename}: {len(all_docs)} chunks (mode={mode_file})")
        print(f"\n🔥 TOTAL JSON CHUNKS: {len(all_docs)}")
        return all_docs
